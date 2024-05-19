from openpyxl import Workbook, load_workbook

# - Load the Excel file -
wb = load_workbook('Example.xlsx')

# - Choose a sheet -
ws = wb['abc']

# - Read data -
print("Example of reading C2: ", ws['C2'].value, "\n\n")

# - Change a value -
ws['C3'].value  = 'Example'

# - Append a new row -
ws.append(['Tim', 2000, '6ft'])

# - Delete last row - 
ws.delete_rows(ws.max_column)

# - Insert an empty row in-between existing ones -
ws.insert_rows(3)

# - Delete a row (same thing with columns) -
ws.delete_rows(3)

# - Iterate a sheet -
print("Sheet iteration:\n")
for row in range(1, 7):
  for column in range(1,4):
    charRepresentation = chr(64 + column)
    print(ws[charRepresentation + str(row)].value)
  print("\n")

# - Get number of columns and rows in a sheet - 
row_count = ws.max_row
column_count = ws.max_column
print("Rows: ", row_count, "Columns: ", column_count)

# - Save changes -
wb.save('Example.xlsx')